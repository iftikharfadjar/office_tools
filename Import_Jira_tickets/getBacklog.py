import argparse
import sys
import openpyxl
import pandas as pd
from openpyxl.utils import range_boundaries
from util import *

def process_excel(input_file, output_file=None, sheet_name=None, keyword="Backlog", target_columns=None, rename_columns=None):
    try:
        # ==========================================
        # STEP 1: OPENPYXL - Unmerge and Fill Cells
        # ==========================================
        print(f"Loading '{input_file}' with openpyxl...")
        wb = openpyxl.load_workbook(input_file, data_only=True)
        
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                print(f"Error: Sheet '{sheet_name}' not found.", file=sys.stderr)
                sys.exit(1)
            ws = wb[sheet_name]
        else:
            ws = wb.active

        print(f"Fixing merged cells in sheet: '{ws.title}'...")

        for merged_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            top_left_cell_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))
            
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = top_left_cell_value

        raw_data = list(ws.values)

        # ==========================================
        # STEP 2: PANDAS - Find Row and Set Headers
        # ==========================================
        print("Converting to pandas DataFrame...")
        df_raw = pd.DataFrame(raw_data)

        mask = df_raw.apply(
            lambda row: row.astype(str).str.contains(keyword, case=False, na=False).any(), 
            axis=1
        )

        if not mask.any():
            print(f"Error: Could not find '{keyword}'.", file=sys.stderr)
            sys.exit(1)

        target_row_index = mask.idxmax()
        
        df_clean = df_raw.iloc[target_row_index:].copy()
        headers = df_clean.iloc[0].astype(str).str.strip()
        df_clean.columns = headers
        df_clean = df_clean[1:] 
        df_clean.reset_index(drop=True, inplace=True)

        end_mask = df_clean.apply(lambda row: row.astype(str).str.contains("TOTAL MANDAYS", case=False, na=False).any(), axis=1)
        if end_mask.any():
            end_index = end_mask.idxmax()
            df_clean = df_clean.iloc[:end_index]

        # ==========================================
        # STEP 3: DYNAMIC COLUMN MAPPING & ORDERING
        # ==========================================
        if target_columns:
            cols_to_keep = [col.strip() for col in target_columns.split(',')]
            existing_cols = [col for col in cols_to_keep if col in df_clean.columns]
            
            if not existing_cols:
                print("Error: None of the requested columns were found.", file=sys.stderr)
                sys.exit(1)
                
            df_clean = df_clean[existing_cols]
        else:
            # df_clean.dropna(axis=1, how='all', inplace=True)
            getting_cols = ['Backlog', 'Acceptance Criteria', 'PIC','Module', 'Mandays']
            df_clean = df_clean[getting_cols]
            df_clean.insert(loc=1, column="Issue Type", value="To Do")
            df_clean.insert(loc=3, column="Priority", value="None")
            df_clean.insert(loc=5, column="Reporter", value="fadjaift")
            df_clean.insert(loc=8, column="Remaining Estimate", value="0")
            df_clean.insert(loc=9, column="Sprint", value="")
            df_clean["Mandays"] = df_clean['Mandays'].apply(rewrite_estimation)
            df_clean["Remaining Estimate"] = df_clean['Mandays']
            df_clean["Module"] = df_clean["Module"].apply(rewrite_components)
            df_clean["Acceptance Criteria"] = df_clean["Acceptance Criteria"].apply(rewrite_description)
            df_clean["PIC"] = df_clean["PIC"].apply(rewrite_assignee)


        # ==========================================
        # STEP 4: RENAMING COLUMNS
        # ==========================================
        if rename_columns:
            rename_dict = {}
            # Split by comma to get pairs, then split by '=' to get old/new names
            for pair in rename_columns.split(','):
                if '=' in pair:
                    old_name, new_name = pair.split('=', 1)
                    rename_dict[old_name.strip()] = new_name.strip()
            
            print(f'Rename_dict = {rename_dict}')
            # Apply the dictionary to rename the columns
            df_clean.rename(columns=rename_dict, inplace=True)
            print(f"Renamed columns: {rename_dict}")
        else:
            rename_dict = {
                'Backlog': 'Summary',
                'Acceptance Criteria' : 'Description',
                'PIC': 'Assignee',
                'Module' : 'Component/s',
                'Mandays' : 'Original Estimate'
            }
            df_clean.rename(columns=rename_dict, inplace=True)
            print(f"Renamed columns: {rename_dict}")
            # df_clean["Original Estimate"] = pd.to_numeric(df_clean['Original Estimate'], errors='coerce') * 8 * 60 * 60

        # ==========================================
        # STEP 5: OUTPUT
        # ==========================================
        if output_file:
            df_clean.to_csv(output_file, index=False, encoding='utf-8')
            print(f"Success! Data saved to '{output_file}'")
        else:
            print("\n--- Final Data Preview ---")
            print(df_clean.head(10).to_string())

    except Exception as e:
        print(f"An unexpected error occurred: {e}", file=sys.stderr)

